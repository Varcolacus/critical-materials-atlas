#!/usr/bin/env python3
"""World Mining Data (Austrian Federal Ministry) -> cube rows. The fourth source.

Prompted by the owner's observation: the site cites seven sources but the cube held three. This
closes the one that genuinely belonged in it.

WMD publishes mine production by country and commodity, and unlike BGS it marks every cell as
REPORTED or ESTIMATED. That flag is the reason to ingest it beyond simple coverage: the cube gains
its first source-declared estimate marker, which the pairing map can use to tell "these two
compilations disagree" apart from "one of them is guessing here".

WHY THE OTHER THREE CITED SOURCES ARE STILL OUT, deliberately and not by omission:
  * IEA - PARTLY SUPERSEDED, see build_cube_iea.py. Its scenario columns stay out for the reason
    below, but the Data Explorer's observed 2024 column is country-level production and IS now
    ingested. The original blanket exclusion was too blunt.
    (Projections are a different grain - scenario x technology x year - and forecasts must not sit
    in a table of observations where a later query could difference them against measured history.)
  * EU CRM 2023 - criticality scores and end-use shares. These are indicators ABOUT a material,
    not observations OF one; they belong as dimensions/attributes, not as rows with a tonnage.
  * ECB reference rates - currency, not mineral. It belongs to the trade-value pipeline.
Each is catalogued rather than ingested, which is what the coverage catalog is for.

Run:  python build_cube_wmd.py     (invoked by build_cube.py; standalone for inspection)
"""
import os, sys, re, csv, glob, warnings
warnings.filterwarnings('ignore')

ROOT = os.environ.get('ATLAS_ROOT', os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'raw', 'wmd', 'wmd_6.4_production_by_country.xlsx')
CODES = os.path.join(ROOT, 'raw', 'baci', 'country_codes_V202601.csv')

# WMD sheet name -> atlas material. The sheet carries the reporting basis in brackets, which is
# information we must keep: "Chromium (Cr2O3)" is an oxide basis, not contained chromium.
SHEET_TO_MATERIAL = {
    'iron': 'iron', 'chromium': 'chromium', 'cobalt': 'cobalt', 'manganese': 'manganese',
    'molybdenum': 'molybdenum', 'nickel': 'nickel', 'niobium': 'niobium', 'tantalum': 'tantalum',
    'titanium': 'titanium', 'tungsten': 'tungsten', 'vanadium': 'vanadium', 'bauxite': 'bauxite',
    'copper': 'copper', 'lead': 'lead', 'zinc': 'zinc', 'tin': 'tin', 'antimony': 'antimony',
    'arsenic': 'arsenic', 'bismuth': 'bismuth', 'cadmium': 'cadmium', 'gold': 'gold',
    'silver': 'silver', 'lithium': 'lithium', 'graphite': 'graphite', 'fluorspar': 'fluorspar',
    'barite': 'baryte', 'baryte': 'baryte', 'magnesite': 'magnesium', 'phosphate': 'phosphate',
    'potash': 'potash', 'salt': 'salt', 'sulfur': 'sulphur', 'gypsum': 'gypsum', 'talc': 'talc',
    'boron': 'boron', 'diamonds': 'diamond', 'mercury': 'mercury', 'selenium': 'selenium',
    'tellurium': 'tellurium', 'zirconium': 'zirconium', 'beryllium': 'beryllium',
    'rare earths': 'rare_earths', 'platinum': 'platinum', 'palladium': 'palladium',
}
# WMD marks each cell: r = reported by the country, e = estimated by WMD.
FLAG = {'r': None, 'e': 'estimated_by_source'}
FLAG_COL = None


def name_to_iso():
    m = {}
    with open(CODES, encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            n, iso = row.get('country_name'), row.get('country_iso3')
            if n and iso and len(iso) == 3:
                m[n.strip().lower()] = iso.strip()
    # WMD spells a handful of countries its own way
    m.update({
        'usa': 'USA', 'united states': 'USA', 'russia': 'RUS', 'south korea': 'KOR',
        'north korea': 'PRK', 'iran': 'IRN', 'vietnam': 'VNM', 'laos': 'LAO', 'syria': 'SYR',
        'tanzania': 'TZA', 'bolivia': 'BOL', 'venezuela': 'VEN', 'moldova': 'MDA',
        'czech republic': 'CZE', 'czechia': 'CZE', 'slovakia': 'SVK', 'turkey': 'TUR',
        'congo, dem. rep.': 'COD', 'dr congo': 'COD', 'd.r. congo': 'COD',
        'congo, d.r.': 'COD', 'congo dr': 'COD', 'bosnia-herzegovina': 'BIH',
        'central african republic': 'CAF', 'dominican republic': 'DOM',
        'french guiana': 'GUF', 'papua new guinea': 'PNG', 'new caledonia': 'NCL',
        'saudi arabia': 'SAU', 'south africa': 'ZAF', 'sierra leone': 'SLE',
        'burkina faso': 'BFA', 'trinidad and tobago': 'TTO', 'sri lanka': 'LKA',
        'congo, rep.': 'COG', 'ivory coast': 'CIV', "cote d'ivoire": 'CIV',
        'united kingdom': 'GBR', 'great britain': 'GBR', 'macedonia': 'MKD',
        'bosnia and herzegovina': 'BIH', 'brunei': 'BRN', 'cape verde': 'CPV',
        'korea, north': 'PRK', 'korea, south': 'KOR', 'kosovo': 'XKX',
        'solomon islands': 'SLB',
    })
    return m


def material_for(sheet):
    base = re.sub(r'\(.*?\)', '', sheet).strip().lower()
    return SHEET_TO_MATERIAL.get(base), (re.search(r'\((.*?)\)', sheet).group(1)
                                         if '(' in sheet else None)


def build():
    import openpyxl
    if not os.path.exists(XLSX):
        return []
    n2i = name_to_iso()
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    rows, unmapped = [], set()
    for sheet in wb.sheetnames:
        material, basis_note = material_for(sheet)
        if not material:
            continue
        ws = wb[sheet]
        years, unit_col = None, None
        for r in ws.iter_rows(max_col=12, values_only=True):     # ONE pass (read_only)
            if r is None or r[0] is None:
                continue
            c0 = str(r[0]).strip()
            if years is None:
                if c0.lower() == 'country':
                    years = {i: int(v) for i, v in enumerate(r)
                             if v is not None and str(v).strip()[:4].isdigit()
                             and 1990 < int(str(v).strip()[:4]) < 2100}
                    unit_col = next((i for i, v in enumerate(r)
                                     if v and str(v).strip().lower() == 'unit'), None)
                    global FLAG_COL
                    FLAG_COL = next((i for i, v in enumerate(r)
                                     if v and 'data source' in str(v).strip().lower()), None)
                continue
            iso = n2i.get(c0.lower())
            if not iso:
                if c0.lower() not in ('total', 'world', 'others', 'other countries'):
                    unmapped.add(c0)
                continue
            unit = (str(r[unit_col]).strip() if unit_col is not None
                    and unit_col < len(r) and r[unit_col] else 'metr. t')
            flag_cell = (str(r[FLAG_COL]).strip().lower()
                         if FLAG_COL is not None and FLAG_COL < len(r) and r[FLAG_COL] else '')
            for i, yr in years.items():
                if i >= len(r) or r[i] in (None, ''):
                    continue
                try:
                    v = float(str(r[i]).replace(',', ''))
                except ValueError:
                    continue
                if v <= 0:
                    continue
                tonnes = v if unit.lower().startswith('metr') else (
                    v / 1000 if unit.lower().startswith('kg') else None)
                rows.append({
                    'material': material, 'source_group': f'WMD:{sheet}', 'country_iso3': iso,
                    'year': yr, 'measure_family': 'production', 'measure': 'production',
                    'flow_direction': None, 'stage': 'mine',
                    'code_system': 'WMD sheet', 'native_code': sheet, 'native_label': sheet,
                    'sub_commodity': basis_note, 'value': v, 'unit': unit,
                    'value_t': tonnes, 'conversion_factor': (1.0 if tonnes == v else
                                                             (0.001 if tonnes else None)),
                    'basis': ('content' if basis_note and any(
                        k in basis_note for k in ('Fe', 'Cr2O3', 'Ti', 'REO', 'content'))
                        else 'gross') if tonnes else None,
                    'source': 'World Mining Data', 'series_id': f'WMD:{sheet}:production',
                    'precision': flag_cell or None, 'value_flag': FLAG.get(flag_cell),
                })
    wb.close()
    if unmapped:
        print(f'  WMD: {len(unmapped)} country names unmapped, e.g. {sorted(unmapped)[:5]}')
    return rows


if __name__ == '__main__':
    import pandas as pd
    sys.stdout.reconfigure(encoding='utf-8')
    df = pd.DataFrame(build())
    if df.empty:
        print('no rows'); raise SystemExit
    print(f'{len(df):,} rows | {df.material.nunique()} materials | '
          f'{df.country_iso3.nunique()} countries | {df.year.min()}-{df.year.max()}')
    print('  estimate-flagged cells:', int((df.value_flag == 'estimated_by_source').sum()))
    print('  materials:', ', '.join(sorted(df.material.unique())[:14]), '...')
