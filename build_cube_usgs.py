#!/usr/bin/env python3
"""USGS Historical Statistics (Data Series 140) -> cube rows.

Found by the coverage catalog: raw/usgs_hist holds 84 workbooks that have been sitting in the repo
with only their price column ever extracted. They carry annual series from 1900 - and, crucially,
66 of them carry a WORLD PRODUCTION column, which the BGS spine (1970-) cannot provide. That makes
this the cheapest available extension of the cube's time depth by seventy years.

What is ingested, and what is deliberately labelled rather than hidden:
  * US series (production, mine/primary/secondary, imports, exports, stocks, shipments) -> USA
  * World production / world mine / world refinery                                      -> WLD
  * Apparent consumption is USGS's OWN derived series. The council's rule is that derived
    analytics do not belong in a fact cube - but a *source-published* derived series is a
    citable object and an independent check on our own apparent-consumption method. It is kept
    under measure_family='derived_by_source' so it can never be mistaken for an observation.
  * Unit values are a different measure family (price) and are kept apart from tonnages. Nominal
    and constant-1998 dollars are distinct measures, not one series.
  * 'W' (withheld) becomes a row with a NULL value and value_flag='withheld' - a company-
    confidential figure is not the same thing as a zero or a gap. 'NA' rows are dropped.

Run:  python build_cube_usgs.py     (invoked by build_cube.py; standalone for inspection)
"""
import os, sys, glob, warnings
warnings.filterwarnings('ignore')

ROOT = os.environ.get('ATLAS_ROOT', os.path.dirname(os.path.abspath(__file__)))
HIST = os.path.join(ROOT, 'raw', 'usgs_hist')

# workbook stem -> atlas material label (only where the atlas has that material; others keep
# their own name and simply ride along as context, exactly as with the BGS groups)
STEM_TO_MATERIAL = {
    'bauxite-and-alumina': 'bauxite', 'phosphate-rock': 'phosphate',
    'platinum-group-metals': 'platinum', 'rare-earths': 'rare_earths',
    'magnesium-metal': 'magnesium', 'magnesium-compounds': 'magnesium',
    'titanium-metal': 'titanium', 'titanium-minerals': 'titanium',
    'titanium-pigments': 'titanium', 'iron-and-steel': 'iron', 'iron-ore': 'iron',
    'barite': 'baryte', 'aluminum': 'aluminium', 'sulfur': 'sulphur',
    'kyanite-and-related-minerals': 'kyanite', 'construction-sand-and-gravel': 'sand_and_gravel',
    'industrial-sand-and-gravel': 'sand_and_gravel', 'boron': 'boron', 'nitrogen': 'nitrogen',
}

# column header (lowercased) -> (measure, measure_family, stage, geography)
COLMAP = {
    'imports': ('imports', 'trade', 'unspecified', 'USA'),
    'imports for consumption': ('imports', 'trade', 'unspecified', 'USA'),
    'exports': ('exports', 'trade', 'unspecified', 'USA'),
    'production': ('production', 'production', 'unspecified', 'USA'),
    'mine production': ('production', 'production', 'mine', 'USA'),
    'primary production': ('production', 'production', 'processed', 'USA'),
    'secondary production': ('production_secondary', 'production', 'processed', 'USA'),
    'smelter production': ('production', 'production', 'processed', 'USA'),
    'refinery production': ('production', 'production', 'processed', 'USA'),
    'shipments': ('shipments', 'trade', 'unspecified', 'USA'),
    'government shipments': ('shipments_government', 'trade', 'unspecified', 'USA'),
    'stocks': ('stocks', 'stocks', 'unspecified', 'USA'),
    'industry stocks': ('stocks_industry', 'stocks', 'unspecified', 'USA'),
    'government stocks': ('stocks_government', 'stocks', 'unspecified', 'USA'),
    'apparent consumption': ('apparent_consumption', 'derived_by_source', 'unspecified', 'USA'),
    'reported consumption': ('consumption', 'consumption', 'unspecified', 'USA'),
    'consumption': ('consumption', 'consumption', 'unspecified', 'USA'),
    'world production': ('production', 'production', 'unspecified', 'WLD'),
    'world mine production': ('production', 'production', 'mine', 'WLD'),
    'world refinery production': ('production', 'production', 'processed', 'WLD'),
    'world production (gross weight)': ('production', 'production', 'unspecified', 'WLD'),
}


def classify(col):
    c = ' '.join(str(col).split()).lower().strip()
    if 'unit value' in c:
        real = '98$' in c or '98 $' in c
        return (('unit_value_real98' if real else 'unit_value_nominal'), 'price',
                'unspecified', 'USA', '1998 USD/t' if real else 'USD/t')
    hit = COLMAP.get(c)
    if hit:
        return hit + ('metric tons',)
    return None


def rows_for(path):
    import openpyxl
    stem = os.path.basename(path)[:-5]
    material = STEM_TO_MATERIAL.get(stem, stem.replace('-', '_'))
    out = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = None
    for row in ws.iter_rows(max_col=20, values_only=True):     # ONE pass (read_only)
        if row is None:
            continue
        c0 = str(row[0]).strip() if row[0] is not None else ''
        if header is None:
            if c0.lower() == 'year':
                header = list(row)
            continue
        if not c0[:4].isdigit():
            continue
        year = int(c0[:4])
        for i, cell in enumerate(row):
            if i == 0 or i >= len(header) or header[i] is None:
                continue
            spec = classify(header[i])
            if not spec:
                continue
            measure, family, stage, geo, unit = spec
            raw = str(cell).strip() if cell is not None else ''
            if raw in ('', 'NA', 'na', '--', 'None'):
                continue
            flag, val = None, None
            if raw.upper() in ('W', 'XX'):
                flag = 'withheld'                     # company-confidential, not a zero
            else:
                try:
                    val = float(raw.replace(',', ''))
                except ValueError:
                    continue
            out.append({
                'material': material, 'source_group': stem, 'country_iso3': geo, 'year': year,
                'measure_family': family, 'measure': measure,
                'flow_direction': {'imports': 'in', 'exports': 'out'}.get(measure),
                'stage': stage, 'code_system': 'USGS DS140 column',
                'native_code': f'{stem}:{str(header[i]).strip()}',
                'native_label': str(header[i]).strip(), 'sub_commodity': None,
                'value': val, 'unit': unit,
                'value_t': val if (unit == 'metric tons' and val is not None) else None,
                'conversion_factor': 1.0 if unit == 'metric tons' else None,
                'basis': 'gross' if unit == 'metric tons' else None,
                'source': 'USGS Historical Statistics (DS 140)',
                'series_id': f'USGS-DS140:{stem}:{measure}:{geo}',
                'precision': None, 'value_flag': flag,
            })
    wb.close()
    return out


def build():
    rows = []
    for f in sorted(glob.glob(os.path.join(HIST, '*.xlsx'))):
        try:
            rows.extend(rows_for(f))
        except Exception as e:
            print(f'  skip {os.path.basename(f)}: {e}')
    return rows


if __name__ == '__main__':
    import pandas as pd
    sys.stdout.reconfigure(encoding='utf-8')
    df = pd.DataFrame(build())
    print(f'{len(df):,} rows from {df.source_group.nunique()} commodities, '
          f'{int(df.year.min())}-{int(df.year.max())}')
    print('  by geography:', df.country_iso3.value_counts().to_dict())
    print('  by family:   ', df.measure_family.value_counts().to_dict())
    print('  withheld:    ', int((df.value_flag == 'withheld').sum()))
