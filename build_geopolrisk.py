#!/usr/bin/env python3
"""
Concentration done right — physical production, the way the field measures supply risk.

The atlas's headline concentration (HHI) is computed on TRADE VALUE, which the engine itself flags as the
most attackable choice: value is nominal, re-export-noisy, and mixes price with quantity. The standard
supply-risk methods — notably the EU Critical Raw Materials global supply-risk factor — instead compute the
HHI on PHYSICAL PRODUCTION shares, weighted by governance (World Bank WGI). This layer does exactly that,
from the precise World Mining Data tonnages. NOTE ON ATTRIBUTION: this is the PRODUCER / global-concentration
view. It is NOT the import-based GeoPolRisk of Gemechu et al. 2016 (Cimprich/Helbig/Sonnemann), whose
defining feature is weighting governance by a specific IMPORTER's actual supplier shares (a consumer-
perspective question). Both are valid; this page answers the producer one. It lays the three concentration
measures side by side so you can see where trade value over- or under-states
real supply concentration.

  HHI_prod,m   = Σ_c (production_share_{c,m})^2            physical mine concentration (0–1)
  gov_risk_m   = Σ_c production_share_{c,m} · grisk_c      production-weighted governance risk, grisk = (2.5 − WGI)/5
  GeoPolRisk_m = HHI_prod,m · gov_risk_m                   concentrated AND badly-governed = worst

Then compare HHI_prod against the atlas's trade-VALUE HHI (data.json) and trade-VOLUME HHI (volume.json).
Public data; deterministic. Run: python build_geopolrisk.py
"""
import json, os
import openpyxl

ROOT = os.path.dirname(os.path.abspath(__file__))
data = {m['label']: m for m in json.load(open(os.path.join(ROOT, 'out', 'data.json'), encoding='utf8'))['materials']}
volume = json.load(open(os.path.join(ROOT, 'out', 'volume.json'), encoding='utf8'))['materials']
WGI = json.load(open(os.path.join(ROOT, 'out', 'wgi.json'), encoding='utf8'))['wgi']
TITLE = {r['label']: r['title'] for r in json.load(open(os.path.join(ROOT, 'out', 'companionality.json'), encoding='utf8'))['rows']}
flows = json.load(open(os.path.join(ROOT, 'out', 'flows_2024.json'), encoding='utf8'))
NAMES = flows.get('names', {})
WMD = os.path.join(ROOT, 'raw', 'wmd', 'wmd_6.4_production_by_country.xlsx')

# atlas label -> WMD sheet (same map as build_production.py)
SHEET = {
    'antimony': 'Antimony', 'arsenic': 'Arsenic', 'baryte': 'Baryte', 'bauxite': 'Bauxite',
    'beryllium': 'Beryllium (conc.)', 'boron': 'Boron Minerals', 'cobalt': 'Cobalt', 'cokingcoal': 'Coking Coal',
    'copper': 'Copper', 'feldspar': 'Feldspar', 'fluorspar': 'Fluorspar', 'gallium': 'Gallium',
    'germanium': 'Germanium', 'graphite': 'Graphite', 'lithium': 'Lithium (Li2O)', 'magnesium': 'Magnesite',
    'magnets': 'Rare Earths (REO)', 'manganese': 'Manganese', 'nickel': 'Nickel', 'niobium': 'Niobium (Nb2O5)',
    'palladium': 'Palladium', 'phosphate': 'Phosphate Rock (P2O5)', 'phosphorus': 'Phosphate Rock (P2O5)',
    'platinum': 'Platinum', 'tantalum': 'Tantalum (Ta2O5)', 'titanium': 'Titanium (TiO2)',
    'tungsten': 'Tungsten (W)', 'vanadium': 'Vanadium (V)',
}
N2I = {  # WMD country name -> ISO2 (for WGI lookup)
 'China': 'CN', 'Australia': 'AU', 'Russia': 'RU', 'USA': 'US', 'United States': 'US', 'Brazil': 'BR',
 'Canada': 'CA', 'Congo, Dem. Rep.': 'CD', 'Congo, D.R.': 'CD', 'Indonesia': 'ID', 'India': 'IN',
 'South Africa': 'ZA', 'Chile': 'CL', 'Peru': 'PE', 'Kazakhstan': 'KZ', 'Mexico': 'MX', 'Turkey': 'TR',
 'Türkiye': 'TR', 'Ukraine': 'UA', 'Vietnam': 'VN', 'Viet Nam': 'VN', 'Myanmar': 'MM', 'Bolivia': 'BO',
 'Argentina': 'AR', 'Zimbabwe': 'ZW', 'Zambia': 'ZM', 'Philippines': 'PH', 'Malaysia': 'MY', 'Thailand': 'TH',
 'Japan': 'JP', 'Korea, Rep.': 'KR', 'Germany': 'DE', 'France': 'FR', 'Spain': 'ES', 'Sweden': 'SE',
 'Finland': 'FI', 'Poland': 'PL', 'Norway': 'NO', 'Morocco': 'MA', 'Jordan': 'JO', 'Saudi Arabia': 'SA',
 'Iran': 'IR', 'Egypt': 'EG', 'Nigeria': 'NG', 'Ghana': 'GH', 'Tanzania': 'TZ', 'Namibia': 'NA',
 'Botswana': 'BW', 'Mozambique': 'MZ', 'Madagascar': 'MG', 'Rwanda': 'RW', 'Burundi': 'BI', 'Laos': 'LA',
 'Mongolia': 'MN', 'Uzbekistan': 'UZ', 'Tajikistan': 'TJ', 'New Caledonia': 'NC', 'Papua New Guinea': 'PG',
 'Guinea': 'GN', 'Jamaica': 'JM', 'Suriname': 'SR', 'Guyana': 'GY', 'Venezuela': 'VE', 'Colombia': 'CO',
 'Cuba': 'CU', 'Greece': 'GR', 'Italy': 'IT', 'Portugal': 'PT', 'Austria': 'AT', 'Czech Republic': 'CZ',
 'Slovakia': 'SK', 'Bulgaria': 'BG', 'Romania': 'RO', 'Serbia': 'RS', 'Bosnia-Herzegovina': 'BA',
 'North Macedonia': 'MK', 'Georgia': 'GE', 'Armenia': 'AM', 'Azerbaijan': 'AZ', 'Pakistan': 'PK',
 'Afghanistan': 'AF', 'Sri Lanka': 'LK', 'United Kingdom': 'GB', 'Ireland': 'IE', 'Estonia': 'EE',
 'Algeria': 'DZ', 'Angola': 'AO', 'Sierra Leone': 'SL', 'Liberia': 'LR', 'Mauritania': 'MR', 'Mali': 'ML',
 'Burkina Faso': 'BF', "Cote d'Ivoire": 'CI', 'Senegal': 'SN', 'Uganda': 'UG', 'Ethiopia': 'ET', 'Kenya': 'KE',
 'United Arab Emirates': 'AE', 'Qatar': 'QA', 'Oman': 'OM', 'Israel': 'IL', 'New Zealand': 'NZ',
 'Dominican Republic': 'DO', 'Panama': 'PA', 'Ecuador': 'EC', 'Gabon': 'GA', 'Cameroon': 'CM', 'Malawi': 'MW',
 'Eritrea': 'ER', 'Nepal': 'NP', 'Bhutan': 'BT', 'Cambodia': 'KH', 'Belgium': 'BE', 'Netherlands': 'NL',
}
WGI_MEDIAN = sorted(WGI.values())[len(WGI) // 2]
def grisk(iso):
    w = WGI.get(iso, WGI_MEDIAN)
    return max(0.0, min(1.0, (2.5 - w) / 5.0))   # 0 = best governed, 1 = worst

wb = openpyxl.load_workbook(WMD, read_only=True, data_only=True)
def parse_sheet(sheet):
    ws = wb[sheet]; rows = list(ws.iter_rows(values_only=True))
    hdr = next((i for i, r in enumerate(rows) if r and str(r[0]).strip() == 'Country'), 1)
    try: c24 = rows[hdr].index('2024')
    except ValueError: c24 = 6
    out = {}
    for r in rows[hdr + 1:]:
        if not r or not r[0]: continue
        name = str(r[0]).strip()
        if name.lower() in ('total', 'world', 'total world', 'others'): continue
        v = r[c24] if c24 < len(r) else None
        if isinstance(v, (int, float)) and v > 0: out[name] = float(v)
    return out

rows = []
for lab, sheet in SHEET.items():
    if sheet not in wb.sheetnames: continue
    prod = parse_sheet(sheet)
    world = sum(prod.values())
    if world <= 0: continue
    shares = {c: t / world for c, t in prod.items()}
    hhi_prod = round(sum(s * s for s in shares.values()), 3)
    gov = round(sum(s * grisk(N2I.get(c)) for c, s in shares.items()), 3)
    geopol = round(hhi_prod * gov, 3)
    cov = round(100 * sum(s for c, s in shares.items() if N2I.get(c) in WGI) / 1.0, 0)  # % of production with a WGI match
    top = max(shares.items(), key=lambda kv: kv[1])
    hhi_val = data.get(lab, {}).get('hhi')
    q = volume.get(lab, {}).get('qty_hhi')
    hhi_vol = q[-1] if q else None
    rows.append({
        'label': lab, 'title': TITLE.get(lab, lab),
        'hhi_prod': hhi_prod, 'gov_risk': gov, 'geopolrisk': geopol, 'wgi_coverage': cov,
        'top_producer': top[0], 'top_iso': N2I.get(top[0]), 'top_share': round(100 * top[1], 1),
        'n_producers': len(shares),
        'hhi_value': hhi_val, 'hhi_volume': hhi_vol,
        'value_minus_prod': (round(hhi_val - hhi_prod, 3) if hhi_val is not None else None),
    })

# rankings & movers: does the trade-value view mis-rank concentration vs the standard production view?
by_prod = sorted(rows, key=lambda r: -r['hhi_prod'])
by_val = sorted([r for r in rows if r['hhi_value'] is not None], key=lambda r: -r['hhi_value'])
prank = {r['label']: i + 1 for i, r in enumerate(by_prod)}
vrank = {r['label']: i + 1 for i, r in enumerate(by_val)}
for r in rows:
    r['rank_prod'] = prank.get(r['label'])
    r['rank_value'] = vrank.get(r['label'])
    r['rank_shift'] = (r['rank_value'] - r['rank_prod']) if (r['rank_value'] and r['rank_prod']) else None

movers = sorted([r for r in rows if r['rank_shift'] is not None], key=lambda r: -abs(r['rank_shift']))[:6]
# where value MOST overstates concentration vs physical production (price/processing effect)
overstated = sorted([r for r in rows if r['value_minus_prod'] is not None], key=lambda r: -r['value_minus_prod'])[:5]
understated = sorted([r for r in rows if r['value_minus_prod'] is not None], key=lambda r: r['value_minus_prod'])[:5]

geopol_rank = sorted(rows, key=lambda r: -r['geopolrisk'])

# ---- IMPORT-BASED GeoPolRisk per bloc: the real Gemechu 2016 consumer view ----
# The score above is the producer/global view. This is the consumer one: weight governance by each bloc's
# ACTUAL bilateral supplier shares (BACI 2023), not global production. Same functional form as the producer
# version for comparability: GeoPolRisk_import(A,M) = HHI(import shares of A) x sum_i share_i * grisk_i.
# An entrepot-stripped variant (drop known trade hubs) is the sensitivity check.
_bi = json.load(open(os.path.join(ROOT, 'raw', 'geopolrisk', 'bilateral_imports_2023.json'), encoding='utf8'))
IMP_BLOCS = ['China', 'EU', 'US', 'Japan', 'Korea', 'India']
_ENTRE = {'NL', 'BE', 'SG', 'HK', 'AE', 'GB', 'CH'}
def _igpr(supp, strip=False):
    d = {s: q for s, q in supp.items() if not (strip and s in _ENTRE)}
    tot = sum(d.values())
    if tot <= 0:
        return None
    sh = {s: q / tot for s, q in d.items()}
    hhi = sum(v * v for v in sh.values())
    gov = sum(v * grisk(s) for s, v in sh.items())
    return round(hhi * gov, 3)
import_rows = []
for r in rows:
    hs = _bi['material_hs'].get(r['label'])
    per = _bi['imports'].get(hs, {}) if hs else {}
    blocs = {b: _igpr(per[b]) for b in IMP_BLOCS if b in per}
    strip = {b: _igpr(per[b], strip=True) for b in IMP_BLOCS if b in per}
    if blocs:
        vals = [v for v in blocs.values() if v is not None]
        import_rows.append({
            'label': r['label'], 'title': r['title'], 'production_gpr': r['geopolrisk'],
            'blocs': blocs, 'blocs_stripped': strip,
            'spread_min': min(vals) if vals else None, 'spread_max': max(vals) if vals else None,
            'shared_hs': len(_bi.get('imports', {}).get(hs, {})) and hs in ('811292',),
        })
# validation gate: (b) reorders vs production for the EU; (c) entrepot-stable
_eu = {r['label']: r['blocs'].get('EU') for r in import_rows if r['blocs'].get('EU') is not None}
_eu_strip = {r['label']: r['blocs_stripped'].get('EU') for r in import_rows if r['blocs_stripped'].get('EU') is not None}
_prank = sorted(_eu, key=lambda l: -(next((rr['production_gpr'] for rr in import_rows if rr['label'] == l), 0) or 0))
_irank = sorted(_eu, key=lambda l: -_eu[l])
_srank = sorted(_eu_strip, key=lambda l: -_eu_strip[l])
import_validation = {
    'eu_reordered': sum(1 for i, l in enumerate(_irank) if _prank.index(l) != i),
    'eu_n': len(_irank),
    'eu_top5_entrepot_stable': len(set(_irank[:5]) & set(_srank[:5])),
    'note': 'Import-based GeoPolRisk (Gemechu 2016 consumer view): governance weighted by each bloc\'s '
            'actual bilateral suppliers (BACI 2023), per metal. HHI(import) x sum share*grisk. The '
            'entrepot-stripped variant (drop NL/BE/CH/GB/SG/HK/AE) is the sensitivity check. Same metal '
            'carries different risk for different blocs - that is the whole point, and the producer view '
            'hides it. Caveat: gallium/germanium/hafnium share HS 811292, so their import shares are '
            'bundled; and BACI records the shipping partner, not always the ultimate origin.',
}

out = {
    'generated': None, 'year': 2024,
    'method': 'HHI on WMD physical production shares x PRODUCTION-weighted governance (WGI). This is a GLOBAL production-concentration supply-risk (the EU-CRM-style producer view). It is NOT the import-based consumer GeoPolRisk of Gemechu 2016, whose innovation is to weight governance by the ACTUAL supplier shares of a specific importer (a distinct, consumer-perspective question).',
    'n': len(rows),
    'geopolrisk_top': [{'title': r['title'], 'geopolrisk': r['geopolrisk'], 'hhi_prod': r['hhi_prod'],
                        'gov_risk': r['gov_risk'], 'top': r['top_producer'], 'top_share': r['top_share']}
                       for r in geopol_rank[:12]],
    'movers': [{'title': r['title'], 'rank_value': r['rank_value'], 'rank_prod': r['rank_prod'],
                'shift': r['rank_shift'], 'hhi_value': r['hhi_value'], 'hhi_prod': r['hhi_prod']} for r in movers],
    'value_overstates': [{'title': r['title'], 'hhi_value': r['hhi_value'], 'hhi_prod': r['hhi_prod'],
                          'gap': r['value_minus_prod']} for r in overstated],
    'value_understates': [{'title': r['title'], 'hhi_value': r['hhi_value'], 'hhi_prod': r['hhi_prod'],
                           'gap': r['value_minus_prod']} for r in understated],
    'import_geopolrisk': sorted(import_rows, key=lambda r: -(r['spread_max'] or 0)),
    'import_blocs': IMP_BLOCS,
    'import_validation': import_validation,
    'rows': sorted(rows, key=lambda r: -r['hhi_prod']),
}
os.makedirs(os.path.join(ROOT, 'out'), exist_ok=True)
json.dump(out, open(os.path.join(ROOT, 'out', 'geopolrisk.json'), 'w', encoding='utf8'), separators=(',', ':'))
print(f"  import GeoPolRisk: EU reordered {import_validation['eu_reordered']}/{import_validation['eu_n']} vs "
      f"production, entrepot top-5 stable {import_validation['eu_top5_entrepot_stable']}/5")
print('wrote out/geopolrisk.json')
print('  GeoPolRisk top 5:', ', '.join(f"{r['title']} {r['geopolrisk']}" for r in geopol_rank[:5]))
print('  biggest value->production rank shifts:', ', '.join(f"{m['title']} {m['rank_shift']:+d}" for m in movers))

HTML = r'''<!doctype html>
<html lang="en"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Concentration, measured properly — production tonnes &amp; GeoPolRisk · Critical Materials Atlas</title>
<meta name="description" content="The atlas's headline concentration uses trade value; the field's standard (EU CRM, GeoPolRisk) uses physical production tonnes weighted by governance. This layer computes both from World Mining Data and shows where trade value misleads.">
<meta property="og:title" content="Concentration measured properly: production tonnes, not trade dollars">
<meta property="og:image" content="https://varcolacus.github.io/critical-materials-atlas/out/share.png">
<link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<link rel="stylesheet" href="assets/site.css"><script src="assets/nav.js" defer></script>
<style>
 .muted{color:#5a6b68;font-size:.86rem}
 #scatter{width:100%;height:440px}
 .stat4{display:grid;grid-template-columns:repeat(4,1fr);gap:.9rem;margin:1.2rem 0}
 @media(max-width:720px){.stat4{grid-template-columns:repeat(2,1fr)}}
 .stat{background:#fff;border:1px solid #e3e9e8;border-left:4px solid #0e7c74;border-radius:10px;padding:.8rem .9rem}
 .stat .v{font-size:1.4rem;font-weight:800;color:#15323a;letter-spacing:-.02em}
 .stat .l{font-size:.76rem;color:#5a6b68;margin-top:.15rem;line-height:1.35}
 table.tidy{width:100%;border-collapse:collapse;font-size:.86rem;margin:.4rem 0}
 table.tidy th,table.tidy td{padding:.4rem .5rem;border-bottom:1px solid #eef1f0;text-align:left}
 table.tidy th.n,table.tidy td.n{text-align:right;font-variant-numeric:tabular-nums}
 .three{display:grid;grid-template-columns:130px 1fr;gap:.5rem;align-items:center;margin:.2rem 0;font-size:.85rem}
 .three .nm{text-align:right;font-weight:600;color:#15323a}
 .bars3{display:flex;flex-direction:column;gap:2px}
 .b3{display:flex;align-items:center;gap:.4rem;font-size:.72rem;color:#5a6b68}
 .b3 .t{height:10px;border-radius:3px}
 .keyline{background:#f2f6f5;border:1px solid #d9e6e3;border-left:4px solid #0e7c74;border-radius:10px;padding:.9rem 1.1rem;margin:1.2rem 0}
 .keyline b{color:#0e7c74}
 .up{color:#c0392b;font-weight:700}.dn{color:#2f8f6b;font-weight:700}
</style>
</head><body>
<header class="topbar"><div class="wrap">
  <a class="wordmark" href="./"><span class="mark"></span>Critical Materials Atlas</a>
  <nav class="topnav"><a href="./">Atlas</a><a href="criticality.html">Criticality</a><a href="production.html">Production</a>
  <a href="volume.html" class="hideable">Value vs volume</a><a href="methodology.html" class="hideable">Methodology</a>
  <a href="https://github.com/Varcolacus/comtrade-reconcile" class="hideable">Engine</a></nav>
</div></header>
<section class="hero"><div class="wrap">
  <div class="eyebrow">Method · concentration · GeoPolRisk</div>
  <h1>Concentration, measured properly</h1>
  <p class="deck">Most of this atlas scores concentration from <i>trade dollars</i> — the number the engine itself calls its most attackable choice. The field&rsquo;s standard (EU Critical Raw Materials, the GeoPolRisk indicator) instead uses <b>physical production tonnes, weighted by governance</b>. This page computes that from the World Mining Data tonnages, and lays value, volume and production side by side to show exactly where the dollar view misleads.</p>
</div></section>
<article style="max-width:1040px">
  <div class="callout"><span id="lead"></span>
  <details class="howto"><summary>The method, and the literature it follows</summary>
  <p>For each material: <b>HHI<sub>prod</sub></b> = sum of squared <i>production</i> shares (World Mining Data 2024 tonnes, all producing countries) — the physical concentration. <b>Governance risk</b> = each producer&rsquo;s share × its World Bank WGI risk (worse-governed producers count more). <b>GeoPolRisk</b> = HHI<sub>prod</sub> × governance risk. We then set HHI<sub>prod</sub> beside the atlas&rsquo;s trade-<b>value</b> HHI and trade-<b>volume</b> HHI.</p>
  <p class="howto-src"><b>What this is, precisely:</b> a <b>global production-concentration</b> supply-risk — physical mine HHI × <i>production-weighted</i> governance — in the spirit of the <a href="https://rmis.jrc.ec.europa.eu/eu-critical-raw-materials" target="_blank" rel="noopener">EU CRM</a> global supply-risk factor. <b>It is not the import-based GeoPolRisk of <a href="https://www.sciencedirect.com/science/article/pii/S0921344924003951" target="_blank" rel="noopener">Gemechu et al. 2016 (Cimprich/Helbig/Sonnemann)</a></b>, whose defining move is to weight governance by a specific importer&rsquo;s <i>actual supplier shares</i> — a consumer-perspective question this page does not answer. Both are legitimate; they are different questions, and this is the producer/global one. <b>Caveats:</b> global, not importer-specific; a few small producers lack a WGI match (coverage shown); phosphorus shares phosphate&rsquo;s sheet. Inputs: WMD tonnes × <a href="out/wgi.json">wgi.json</a> × <a href="out/data.json">data.json</a> &rarr; <a href="out/geopolrisk.json">geopolrisk.json</a>.</p>
  </details></div>

  <div class="stat4" id="stats"></div>
  <div class="keyline" id="keyline"></div>

  <h2 style="margin:1.6rem 0 .3rem">Three ways to measure the same concentration</h2>
  <p class="muted" style="margin-top:0">Each dot is a material: trade-<b>value</b> HHI (x, what the atlas headlines) vs <b>production</b> HHI (y, the standard). On the line = they agree. <span style="color:#c0392b;font-weight:700">Below</span> = trade value <i>overstates</i> concentration (a price or processing effect); <span style="color:#2f8f6b;font-weight:700">above</span> = it <i>understates</i> it.</p>
  <div id="scatter"></div>

  <h2 style="margin:1.6rem 0 .3rem">Supply risk on the standard measure (GeoPolRisk)</h2>
  <p class="muted" style="margin-top:0">Production concentration × governance risk — the metals where a lot of world supply sits in few, and badly-governed, hands.</p>
  <table class="tidy" id="gtab"><thead><tr><th>Material</th><th class="n">GeoPolRisk</th><th class="n">production HHI</th><th class="n">governance risk</th><th>top producer</th></tr></thead><tbody></tbody></table>

  <h2 style="margin:1.8rem 0 .3rem">Every material — value vs volume vs production</h2>
  <table class="tidy" id="ctab"><thead><tr><th>Material</th><th>concentration (0–1): value · volume · production</th><th class="n">value→prod rank shift</th></tr></thead><tbody></tbody></table>

  <h2 style="margin:1.8rem 0 .3rem">The consumer view: who <i>you</i> buy from (import-based GeoPolRisk)</h2>
  <p class="muted" style="margin-top:0">Everything above is the <b>producer</b> view — how concentrated and badly-governed is <i>global</i> supply. The real Gemechu-2016 GeoPolRisk asks a different question: weight governance by each bloc&rsquo;s <b>actual bilateral suppliers</b> (who it imports from), not global production. The same metal then carries very different risk for different blocs.</p>
  <div class="keyline" id="importlead"></div>
  <div style="overflow-x:auto"><table class="tidy" id="imptab"><thead></thead><tbody></tbody></table></div>
  <p class="muted" id="impgate" style="margin-top:.5rem"></p>

  <h2 style="margin:1.8rem 0 .3rem">Why this matters</h2>
  <p>This puts the atlas on the same footing as the EU&rsquo;s official criticality work and the GeoPolRisk literature: supply concentration measured where it physically happens — the mine — not where the money changes hands. Where the two agree, the trade-value story was safe; where they diverge, the dollar view was quietly reporting a price or a refining hub as if it were a mine. It doesn&rsquo;t replace the trade lens — trade is still how material actually moves and where chokepoints bite — but it anchors the concentration claim in tonnes, and closes the gap the engine&rsquo;s own caveat left open.</p>
</article>
<footer class="siteftr"><div class="wrap">
  <div><h4>Critical Materials Atlas</h4>An independent demonstration from public data. Not affiliated with, nor representing, any institution.</div>
  <div><h4>Navigate</h4><a href="criticality.html">Criticality</a><br><a href="production.html">Production in tonnes</a><br><a href="volume.html">Value vs volume</a><br><a href="methodology.html">Methodology</a></div>
  <div><h4>Sources</h4>World Mining Data 2026 (production tonnes) · World Bank WGI · trade HHI — GeoPolRisk / EU-CRM aligned</div>
  <div class="fineprint">A global production-concentration supply-risk (EU-CRM-style producer view), not the import-based consumer GeoPolRisk of Gemechu 2016; governance from WGI.</div>
</div></footer>
<script>
function ld(u){return new Promise((res,rej)=>{const s=document.createElement('script');s.src=u;s.onload=res;s.onerror=rej;document.head.appendChild(s);});}
Promise.all([fetch('out/geopolrisk.json').then(r=>r.json()),
  ld('https://cdn.jsdelivr.net/npm/echarts@5/dist/echarts.min.js')]).then(([S])=>{
  const ov=S.value_overstates[0];
  document.getElementById('lead').innerHTML='<b>Result:</b> measured the standard way — physical production tonnes, not trade dollars — the concentration ranking shifts for several materials. Trade value most <i>overstates</i> concentration for <b>'+S.value_overstates.slice(0,2).map(x=>x.title).join(' and ')+'</b> (a price or refining effect), and the highest supply risk once governance is folded in sits with '+S.geopolrisk_top.slice(0,3).map(x=>x.title).join(', ')+'. The physical measure is the one the EU and the GeoPolRisk literature actually use.';
  // ---- import-based GeoPolRisk (consumer view) ----
  var IG=S.import_geopolrisk, IB=S.import_blocs, IV=S.import_validation;
  if(IG && IG.length){
    var ga=IG.find(function(r){return r.label==='gallium';});
    document.getElementById('importlead').innerHTML='<b>The producer view can mislead a specific bloc.</b> '+(ga?'Gallium is the scariest metal by global production ('+ga.production_gpr+', 97% China) &mdash; yet on <i>actual imports</i> it is <b>low-risk for the EU</b> ('+(ga.blocs.EU*1000).toFixed(0)+') and <b>high for Korea</b> ('+(ga.blocs.Korea*1000).toFixed(0)+'): they buy from different places. ':'')+'Weighting governance by who each bloc actually buys from reorders <b>'+IV.eu_reordered+' of '+IV.eu_n+'</b> metals for the EU vs the producer ranking, and the top-5 survives stripping trade hubs ('+IV.eu_top5_entrepot_stable+'/5). It is a different, complementary question &mdash; not a replacement.';
    var th='<tr><th>Metal</th><th class="n">producer</th>'+IB.map(function(b){return '<th class="n">'+b+'</th>';}).join('')+'</tr>';
    document.querySelector('#imptab thead').innerHTML=th;
    var body=IG.slice(0,12).map(function(r){
      var cells=IB.map(function(b){var v=r.blocs[b]; if(v==null) return '<td class="n" style="color:#c9d2d0">–</td>';
        var col=v>=0.35?'#c0392b':v>=0.15?'#d98324':'#5a6b68';
        return '<td class="n" style="color:'+col+';font-weight:'+(v>=0.35?700:400)+'">'+(v*1000).toFixed(0)+'</td>';}).join('');
      return '<tr><td><b>'+r.title+'</b>'+(r.shared_hs?' <span title="shares HS 811292 with Ge/Hf" style="color:#b07a18">⛓</span>':'')+'</td><td class="n" style="color:#5a6b68">'+(r.production_gpr*1000).toFixed(0)+'</td>'+cells+'</tr>';
    }).join('');
    document.querySelector('#imptab tbody').innerHTML=body;
    document.getElementById('impgate').innerHTML='Scores ×1000; red ≥350, amber ≥150. '+IV.note;
  }
  const stats=[
    {v:S.geopolrisk_top[0].title,l:'highest GeoPolRisk (production concentration × governance)'},
    {v:'±'+Math.max.apply(null,S.movers.map(m=>Math.abs(m.shift))),l:'biggest rank change between the value and production views ('+S.movers[0].title+')'},
    {v:S.value_overstates[0].title,l:'where trade value most overstates real concentration'},
    {v:S.n,l:'materials measured in physical production tonnes'},
  ];
  document.getElementById('stats').innerHTML=stats.map(s=>'<div class="stat"><div class="v">'+s.v+'</div><div class="l">'+s.l+'</div></div>').join('');
  document.getElementById('keyline').innerHTML='<b>The point:</b> a Herfindahl on trade <i>value</i> can call a material “concentrated” when what’s really concentrated is its <i>price</i> or its <i>refining</i>. Recomputing on production <i>tonnes</i> — the EU-CRM / GeoPolRisk standard — separates the two. It is the direct answer to the engine’s own admission that trade value is its most attackable input.';
  // scatter value vs prod
  const pts=S.rows.filter(r=>r.hhi_value!=null).map(r=>{
    const over=(r.hhi_value-r.hhi_prod);
    return {value:[r.hhi_value,r.hhi_prod,r.title],itemStyle:{color:(over>0.08?'#c0392b':over<-0.08?'#2f8f6b':'#8a9aa0')+'cc'},symbolSize:10};
  });
  const ch=echarts.init(document.getElementById('scatter'));
  ch.setOption({grid:{left:52,right:24,top:20,bottom:48},
    tooltip:{formatter:p=>'<b>'+p.value[2]+'</b><br>value HHI: '+p.value[0]+'<br>production HHI: '+p.value[1]},
    xAxis:{name:'trade-value HHI',nameLocation:'middle',nameGap:30,min:0,max:1,axisLabel:{color:'#5a6b68'},nameTextStyle:{color:'#5a6b68'},splitLine:{lineStyle:{color:'#eef1f0'}}},
    yAxis:{name:'production HHI (standard)',nameLocation:'middle',nameGap:38,min:0,max:1,axisLabel:{color:'#5a6b68'},nameTextStyle:{color:'#5a6b68'},splitLine:{lineStyle:{color:'#eef1f0'}}},
    series:[{type:'scatter',data:pts,label:{show:true,formatter:p=>p.value[2],position:'right',fontSize:9,color:'#15323a',distance:4},
      markLine:{silent:true,symbol:'none',lineStyle:{color:'#c9b3ad',type:'dashed'},data:[[{coord:[0,0]},{coord:[1,1]}]]}}]});
  window.addEventListener('resize',()=>ch.resize());
  // geopolrisk table
  const gt=document.querySelector('#gtab tbody');
  S.geopolrisk_top.forEach(r=>{const tr=document.createElement('tr');
    tr.innerHTML='<td><b>'+r.title+'</b></td><td class="n" style="font-weight:700;color:'+(r.geopolrisk>=0.4?'#c0392b':r.geopolrisk>=0.2?'#b07a18':'#5a6b68')+'">'+r.geopolrisk.toFixed(2)+'</td>'+
      '<td class="n">'+r.hhi_prod.toFixed(2)+'</td><td class="n">'+r.gov_risk.toFixed(2)+'</td><td class="muted">'+r.top+' ('+r.top_share+'%)</td>';
    gt.appendChild(tr);});
  // three-way bars table
  const ct=document.querySelector('#ctab tbody');
  const bar=(v,col)=>v==null?'<span style="color:#c9d2d0">–</span>':'<span class="b3"><span class="t" style="width:'+Math.max(2,v*120)+'px;background:'+col+'"></span>'+v.toFixed(2)+'</span>';
  S.rows.forEach(r=>{
    const sh=r.rank_shift;
    const shc=sh==null?'–':(sh>0?'<span class="up">▲ '+sh+'</span>':sh<0?'<span class="dn">▼ '+(-sh)+'</span>':'—');
    const tr=document.createElement('tr');
    tr.innerHTML='<td><b>'+r.title+'</b></td><td><div class="bars3">'+
      bar(r.hhi_value,'#b07a18')+bar(r.hhi_volume,'#3f6fb0')+bar(r.hhi_prod,'#2f8f6b')+'</div></td>'+
      '<td class="n">'+shc+'</td>';
    ct.appendChild(tr);
  });
});
</script>
</body></html>'''
open(os.path.join(ROOT, 'geopolrisk.html'), 'w', encoding='utf8', newline='\n').write(HTML)
print('wrote geopolrisk.html')
