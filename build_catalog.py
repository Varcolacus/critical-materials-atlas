#!/usr/bin/env python3
"""THE COVERAGE CATALOG — what data we could use, before we pay to use it.

The owner's question: "can the cube be bigger? Having a rich cube over our scope lets us check
quickly and cheaply whether data exists when a question arises, and whether a new sub-project is
even possible."

The council's answer (Codex): the *idea* is right but full ingestion is a maintenance trap. Build a
catalog of SERIES, not a bigger table of observations. One row per available series with its
geography, year span and parse status. Then feasibility is a lookup, and only series that earn
their place get promoted into cube.parquet.

So this builder answers "do we have data for X?" without normalizing a single extra value.

    status vocabulary
      in_cube   — already ingested and queryable in pipeline/data/cube.parquet
      on_disk   — the file is downloaded and sitting in raw/, but not parsed into the cube
      reachable — a known public series from an institution we already use, not yet fetched
                  (marked unverified: the URL/scope has NOT been re-checked by this script)
      declined  — considered and rejected, with the reason. A catalog that only lists what we
                  might use is half a record: free, well-licensed and adjacent is not the same as
                  relevant, and an unwritten refusal gets re-proposed every few months.

Run:  python build_catalog.py   ->  out/catalog.json + pipeline/data/catalog.parquet
"""
import os, sys, json, glob, zipfile
import pandas as pd

sys.stdout.reconfigure(encoding='utf-8')
ROOT = os.environ.get('ATLAS_ROOT', os.path.dirname(os.path.abspath(__file__)))
rows = []


def add(**kw):
    rows.append({
        'institution': kw.get('institution'), 'dataset': kw.get('dataset'),
        'series': kw.get('series'), 'measure_family': kw.get('measure_family'),
        'geography': kw.get('geography'), 'year_min': kw.get('year_min'),
        'year_max': kw.get('year_max'), 'n_countries': kw.get('n_countries'),
        'n_rows': kw.get('n_rows'), 'unit': kw.get('unit'), 'frequency': kw.get('frequency', 'annual'),
        'status': kw.get('status'), 'path': kw.get('path'), 'note': kw.get('note'),
    })


# ── 1. what is already IN THE CUBE ────────────────────────────────────────────────────────────
cube_path = os.path.join(ROOT, 'pipeline', 'data', 'cube.parquet')
if os.path.exists(cube_path):
    c = pd.read_parquet(cube_path)
    g = c.groupby(['source', 'material', 'measure'])
    for (src, mat, meas), d in g:
        add(institution='BGS', dataset='World Mineral Statistics', series=f'{mat} · {meas}',
            measure_family='trade' if meas in ('imports', 'exports') else 'production',
            geography='global', year_min=int(d.year.min()), year_max=int(d.year.max()),
            n_countries=int(d.country_iso3.nunique()), n_rows=int(len(d)),
            unit='tonnes (harmonized)', status='in_cube', path='pipeline/data/cube.parquet')

# ── 2. ON DISK but not parsed: USGS Historical Statistics (the big one) ────────────────────────
# 84 workbooks of US series running from ~1900 — mine/smelter/primary/secondary production,
# shipments, imports, exports. Only the price column was ever extracted.
try:
    import openpyxl
    for f in sorted(glob.glob(os.path.join(ROOT, 'raw', 'usgs_hist', '*.xlsx'))):
        commodity = os.path.basename(f)[:-5]
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            header, years = None, []
            for row in ws.iter_rows(max_col=14, values_only=True):   # ONE pass only (read_only)
                if row is None:
                    continue
                first = str(row[0]).strip() if row[0] is not None else ''
                if header is None and first.lower() == 'year':
                    header = [str(x).strip() for x in row[1:] if x is not None]
                    continue
                if header is not None and first[:4].isdigit():
                    years.append(int(first[:4]))
            wb.close()
            if header and years:
                add(institution='USGS', dataset='Historical Statistics (DS 140)',
                    series=f'{commodity} · ' + ', '.join(header[:5]) + ('…' if len(header) > 5 else ''),
                    measure_family='production+trade+price', geography='United States',
                    year_min=min(years), year_max=max(years), n_countries=1, n_rows=len(years),
                    unit='varies (metric tons unless noted)', status='on_disk',
                    path=os.path.relpath(f, ROOT),
                    note=f'{len(header)} measures; only the price column has ever been extracted')
        except Exception as e:
            add(institution='USGS', dataset='Historical Statistics (DS 140)', series=commodity,
                measure_family='production+trade+price', geography='United States',
                status='on_disk', path=os.path.relpath(f, ROOT), note=f'unparsed: {e}')
except ImportError:
    pass

# ── 3. ON DISK, partially used: BACI trade vintages ────────────────────────────────────────────
for z in sorted(glob.glob(os.path.join(ROOT, 'raw', 'baci', '*.zip'))):
    try:
        names = zipfile.ZipFile(z).namelist()
        yrs = sorted({int(n.split('_Y')[1][:4]) for n in names if '_Y' in n})
        vint = os.path.basename(z).replace('BACI_', '').replace('.zip', '')
        add(institution='CEPII', dataset=f'BACI {vint}', series='bilateral trade, all HS6',
            measure_family='trade', geography='global (bilateral)', year_min=min(yrs),
            year_max=max(yrs), n_rows=len(yrs), unit='tonnes + USD', status='on_disk',
            path=os.path.relpath(z, ROOT),
            note='not yet in the cube; HS17 used for the snapshot layer, HS02 for trend pages')
    except Exception:
        pass

# ── 4. ON DISK: other holdings worth surfacing ────────────────────────────────────────────────
for pat, inst, ds, fam, geo, note in [
    ('raw/wmd/*.xlsx', 'World Mining Data', 'Production by country', 'production', 'global',
     'used for the 3-source production cross-check; not in the cube'),
    ('raw/usgs_mcs/*.pdf', 'USGS', 'Mineral Commodity Summaries', 'production+reserves', 'global',
     'reserves, refinery output, import reliance and recycling largely unextracted'),
    ('raw/usgs_critmin/*.csv', 'USGS', 'Critical minerals map (PP1802)', 'deposits', 'global',
     'deposit points; no time dimension'),
    ('raw/bgs/panel/*.json', 'BGS', 'World Mineral Statistics (raw pull)', 'production+trade',
     'global', 'fully ingested into the cube'),
]:
    fs = glob.glob(os.path.join(ROOT, pat))
    if fs:
        add(institution=inst, dataset=ds, series=f'{len(fs)} files', measure_family=fam,
            geography=geo, n_rows=len(fs), status='on_disk', path=pat, note=note)

# ── 5. REACHABLE but not fetched — same institutions, unexploited series ──────────────────────
# Named by the council. Deliberately NOT given URLs: nothing here has been re-verified by this
# script, and an unchecked URL on a feasibility sheet is exactly the kind of thing that later gets
# quoted as if it had been checked.
for inst, ds, fam, geo, note in [
    ('BGS', 'European Mineral Statistics', 'production+trade', 'Europe', 'companion to WMS'),
    ('BGS', 'Risk List indicators', 'risk', 'global', 'supply-risk scores per commodity'),
    ('USGS', 'Minerals Yearbook (international)', 'production+capacity', 'global',
     'country chapters; richer than MCS but PDF-bound'),
    ('CEPII', 'BACI HS92 vintage', 'trade', 'global (bilateral)',
     'would extend trade back to 1995 with code concordance'),
    ('IEA', 'Critical Minerals Market Review tables', 'demand+capacity', 'global',
     'scenario demand; keep forecasts separate from observations'),
    ('EU JRC', 'RMIS / CRM assessment indicators', 'risk+end-use', 'EU',
     'supply risk, economic importance, end-use shares, substitution'),
    ('Eurostat', 'PRODCOM', 'production', 'EU', 'industrial production by product code'),
    ('Eurostat', 'COMEXT (full)', 'trade', 'EU', 'CN8 detail; partially held'),
    ('Eurostat', 'Circular material use / waste', 'recycling', 'EU', 'secondary-supply angle'),
    ('Geoscience Australia', 'Australian mineral resources', 'reserves+production', 'Australia', ''),
    ('NRCan', 'Canadian minerals yearbook', 'reserves+production', 'Canada', ''),
    ('BGR (DERA)', 'German raw-materials data', 'production+risk', 'global', ''),
]:
    add(institution=inst, dataset=ds, series='(whole dataset)', measure_family=fam, geography=geo,
        status='reachable', note=(note + ' — scope/URL NOT verified by this build').strip(' —'))

# ── 6. CONSIDERED AND DECLINED ────────────────────────────────────────────────────────────────
# A catalog that only lists what we might use is half a record. Free, well-licensed and adjacent is
# not the same as relevant, and without writing the refusal down the same dataset gets re-proposed
# every few months. Each entry names what it is and why it does not belong here.
for inst, ds, why in [
    ('IEA', 'Building-level Electricity Access and Demand Model (BEACON / LItLDF)',
     'Out of scope, not out of quality. It estimates electricity ACCESS and DEMAND per BUILDING in '
     'sub-Saharan Africa from satellite imagery - a different subject, a different unit of '
     'observation and a different geography from mineral production and trade. The only route to '
     'relevance would be building-level demand -> grid buildout -> conductor tonnage, which is a '
     'speculative modelling chain we do not have and could not defend. Where grid expansion is '
     'genuinely needed as a demand driver, IEA network investment series are the direct measure. '
     'CC BY 4.0; code at github.com/stephenjlee/beacon and /litldf if that ever changes.'),
    ('IEA', 'CCUS Projects Database (2026 edition)',
     'INSPECTED, not judged from the description: 2,034 projects, 30 columns, and NOT ONE material '
     'dimension - project name, country, partners, dates, status, CO2 capacity in Mt/yr, sector, '
     'fate of carbon, references. Two independent reasons it cannot enter the cube. (a) Its unit of '
     'observation is a PROJECT; the cube is country-year. (b) Getting from CO2 capacity to material '
     'demand needs tonnes of steel/alloy per Mt of capture capacity, which nobody publishes - and '
     'our consumption model calibrates intensities by back-solving from a KNOWN WORLD TOTAL, of '
     'which there is none for "material embodied in CCUS". We would be inventing the number we then '
     'reported. If an energy-infrastructure materials layer is ever built on published intensities, '
     'this is a good input to it; until then it is a different subject. CC BY 4.0.'),
    ('IEA', 'Demand projections (STEPS / APS / NZE scenarios)',
     'Forecasts. They must not sit in a table of observations where a later query could difference '
     'them against measured history and call the result a trend. Used on the site as cited '
     'projections, never as cube rows.'),
    ('Commercial', 'S&P / Wood Mackenzie / Benchmark asset-level data',
     'Not public and not reproducible. The atlas rests on sources a reader can fetch and check.'),
]:
    add(institution=inst, dataset=ds, series='(whole dataset)', measure_family='n/a',
        geography='n/a', status='declined', note=why)


if __name__ == '__main__':
    df = pd.DataFrame(rows)
    os.makedirs(os.path.join(ROOT, 'pipeline', 'data'), exist_ok=True)
    df.to_parquet(os.path.join(ROOT, 'pipeline', 'data', 'catalog.parquet'), index=False)

    by_status = df.status.value_counts().to_dict()
    summary = {
        'note': 'Coverage catalog: one row per SERIES (not per observation). Answers "do we have '
                'data for X?" before any ingestion work. status: in_cube / on_disk / reachable. '
                '"reachable" rows are named from institutions we already use and are NOT verified '
                'by this build — treat as leads, not facts.',
        'series': int(len(df)),
        'by_status': {k: int(v) for k, v in by_status.items()},
        'institutions': sorted(df.institution.dropna().unique().tolist()),
        'earliest_year_on_disk': int(df.year_min.min()) if df.year_min.notna().any() else None,
        'headline_gap': 'USGS Historical Statistics: 84 workbooks of US series from ~1900 sitting '
                        'in raw/, of which only the price column was ever extracted.',
    }
    json.dump({'summary': summary, 'series': df.where(df.notna(), None).to_dict('records')},
              open(os.path.join(ROOT, 'out', 'catalog.json'), 'w', encoding='utf-8'), indent=1)

    print(f'WROTE catalog — {len(df)} series')
    for k, v in by_status.items():
        print(f'   {k}: {v}')
    print(f'   institutions: {", ".join(summary["institutions"])}')
    print(f'   earliest year held on disk: {summary["earliest_year_on_disk"]}')
